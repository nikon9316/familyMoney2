from aiohttp import web, ClientSession
import logging
from pathlib import Path
import hashlib, hmac, json
from urllib.parse import parse_qsl
from time import time
from datetime import datetime, date
from decimal import Decimal
from io import BytesIO
import tempfile
import os
import secrets
from collections import defaultdict, deque
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy import text
from config import BOT_TOKEN, ADMIN_PANEL_TOKEN, ADMIN_ID, SUPERADMIN_TOKEN, SUPERADMIN_TELEGRAM_ID, ADMIN_IP_ALLOWLIST, ADMIN_RATE_LIMIT_COUNT, ADMIN_RATE_LIMIT_WINDOW, RESTORE_OTP_TTL_SECONDS, ADMIN_SESSION_TTL_SECONDS, ADMIN_CSRF_HEADER, SESSION_COOKIE_SECURE, BUDGET_NOTIFY_COOLDOWN_HOURS, PDF_REPORT_FONT, PDF_REPORT_FONT_PATH, PDF_REPORT_FONT_BOLD_PATH, ERROR_NOTIFY_TELEGRAM, APP_VERSION, HEALTHCHECK_SECRET, DATABASE_URL
from database.db import *
from database.schema import engine
from backup import create_backup, cleanup_old_backups, list_backups, get_backup_path, restore_backup_from_file

BASE_DIR = Path(__file__).resolve().parent
WEBAPP_DIR = BASE_DIR / 'webapp'
ADMIN_DIR = BASE_DIR / 'admin'
logger = logging.getLogger('family_finance.server')
AUTH_MAX_AGE_SECONDS = 60 * 60 * 24


def _capture_exception(exc: BaseException) -> None:
    """Best-effort error capture for Sentry/compatible SDK."""
    try:
        import sentry_sdk
        sentry_sdk.capture_exception(exc)
    except Exception:
        pass

_admin_rate_buckets = defaultdict(deque)
_restore_otps = {}
_admin_sessions = {}
_restore_maintenance_until = 0
_idempotency_cache = {}
_idempotency_inflight = set()
IDEMPOTENCY_TTL_SECONDS = 90

# Level 4.0.1: safe JSON serialization for Decimal/datetime nested inside API payloads.
def _json_default(obj):
    if isinstance(obj, Decimal):
        return float(obj)
    if isinstance(obj, (datetime, date)):
        return obj.isoformat()
    raise TypeError(f'Object of type {type(obj).__name__} is not JSON serializable')

_orig_json_response = web.json_response
def _safe_json_response(data=None, *args, **kwargs):
    kwargs.setdefault('dumps', lambda value: json.dumps(value, default=_json_default, ensure_ascii=False))
    return _orig_json_response(data, *args, **kwargs)
web.json_response = _safe_json_response

def _client_ip(request: web.Request) -> str:
    forwarded = request.headers.get('X-Forwarded-For', '')
    if forwarded:
        return forwarded.split(',')[0].strip()
    peer = request.transport.get_extra_info('peername') if request.transport else None
    return str(peer[0]) if peer else 'unknown'

def _admin_ip_allowed(request: web.Request) -> bool:
    if not ADMIN_IP_ALLOWLIST:
        return True
    return _client_ip(request) in ADMIN_IP_ALLOWLIST

def _rate_limit_admin(request: web.Request):
    ip = _client_ip(request)
    now = time()
    bucket = _admin_rate_buckets[ip]
    while bucket and now - bucket[0] > ADMIN_RATE_LIMIT_WINDOW:
        bucket.popleft()
    if len(bucket) >= ADMIN_RATE_LIMIT_COUNT:
        raise web.HTTPTooManyRequests(text='Too Many Requests')
    bucket.append(now)

@web.middleware
async def admin_security_middleware(request, handler):
    if request.path.startswith('/admin') or request.path.startswith('/api/admin'):
        _rate_limit_admin(request)
        if not _admin_ip_allowed(request):
            try:
                admin_log('blocked_ip', admin_label='unknown', ip_address=_client_ip(request), details=request.path)
            except Exception:
                pass
            raise web.HTTPForbidden(text='Forbidden IP')
    return await handler(request)


@web.middleware
async def idempotency_middleware(request, handler):
    """Protect money-changing POST/PUT/DELETE requests from double taps/retries.

    Client sends X-Idempotency-Key. The server returns the cached JSON result for
    repeated requests with the same Telegram initData + path + method + key.
    """
    if request.method not in {'POST', 'PUT', 'DELETE'} or not request.path.startswith('/api/'):
        return await handler(request)
    key = request.headers.get('X-Idempotency-Key')
    if not key:
        return await handler(request)
    now = time()
    for k, v in list(_idempotency_cache.items()):
        if now - v.get('created_at', 0) > IDEMPOTENCY_TTL_SECONDS:
            _idempotency_cache.pop(k, None)
    user_part = request.headers.get('X-Telegram-Init-Data', '')[:128] or request.remote or 'anonymous'
    cache_key = hashlib.sha256(f'{user_part}|{request.method}|{request.path}|{key}'.encode()).hexdigest()
    cached = _idempotency_cache.get(cache_key)
    if cached:
        return web.json_response(cached['data'], status=cached.get('status', 200))
    if cache_key in _idempotency_inflight:
        return json_error('Запрос уже выполняется. Подождите несколько секунд.', 409)
    _idempotency_inflight.add(cache_key)
    try:
        response = await handler(request)
        body = getattr(response, 'body', None)
        if body and response.content_type == 'application/json':
            try:
                _idempotency_cache[cache_key] = {'created_at': now, 'status': response.status, 'data': json.loads(body.decode())}
            except Exception:
                pass
        return response
    finally:
        _idempotency_inflight.discard(cache_key)


@web.middleware
async def maintenance_middleware(request, handler):
    """Level 5.2: block user API actions while restore maintenance is active.

    Admin restore endpoints and public static/health pages stay available. This
    prevents users from writing financial data while a restore is in progress.
    """
    if time() <= _restore_maintenance_until:
        allowed_prefixes = (
            '/api/admin/restore',
            '/api/admin/login',
            '/api/admin/logout',
            '/api/admin/stats',
            '/admin',
            '/webapp/',
        )
        allowed_paths = {'/', '/healthz', '/privacy'}
        if request.path not in allowed_paths and not any(request.path.startswith(p) for p in allowed_prefixes):
            return web.json_response({
                'ok': False,
                'error': 'Сервис временно в режиме обслуживания: идет восстановление backup. Попробуйте позже.',
                'maintenance_mode': True,
            }, status=503)
    return await handler(request)


def verify_telegram_init_data(init_data: str) -> dict:
    if not init_data:
        raise ValueError('Откройте приложение через Telegram')
    parsed = dict(parse_qsl(init_data, keep_blank_values=True))
    received_hash = parsed.pop('hash', None)
    if not received_hash:
        raise ValueError('Нет подписи Telegram')
    data_check_string = '\n'.join(f'{k}={v}' for k, v in sorted(parsed.items()))
    secret_key = hmac.new(b'WebAppData', BOT_TOKEN.encode(), hashlib.sha256).digest()
    calculated_hash = hmac.new(secret_key, data_check_string.encode(), hashlib.sha256).hexdigest()
    if not hmac.compare_digest(calculated_hash, received_hash):
        raise ValueError('Неверная подпись Telegram')
    auth_date = int(parsed.get('auth_date', '0'))
    if auth_date and time() - auth_date > AUTH_MAX_AGE_SECONDS:
        raise ValueError('Сессия устарела. Откройте WebApp заново')
    if not parsed.get('user'):
        raise ValueError('Telegram не передал пользователя')
    return json.loads(parsed['user'])

def auth_user(request: web.Request) -> dict:
    tg_user = verify_telegram_init_data(request.headers.get('X-Telegram-Init-Data', ''))
    full_name = ' '.join(filter(None, [tg_user.get('first_name',''), tg_user.get('last_name','')])).strip() or tg_user.get('username') or 'Пользователь'
    user = get_or_create_user(int(tg_user['id']), full_name)
    if is_user_blocked(user):
        raise ValueError('Пользователь заблокирован администратором')
    return user

def fid(user):
    return int(user['family_id'])

def json_error(message: str, status=400):
    return web.json_response({'ok': False, 'error': message}, status=status)

async def index(request):
    return web.FileResponse(WEBAPP_DIR / 'index.html')

def init_payload(user: dict):
    family_id = fid(user)
    return {
        'ok': True,
        'user': user,
        'permissions': get_effective_permissions(user),
        'family': get_family(family_id),
        'members': get_members(family_id),
        'summary': get_summary(family_id),
        'month_summary': get_month_summary(family_id),
        'category_report': get_expense_by_categories(family_id),
        'daily_chart': get_daily_chart(family_id),
        'chart_pack': get_webapp_chart_pack(family_id),
        'forecast': get_expense_forecast(family_id),
        'wallets': get_wallets(family_id),
        'income_categories': get_categories(family_id, 'income'),
        'expense_categories': get_categories(family_id, 'expense'),
        'rates': get_rates(family_id),
        'debts': get_debts(family_id),
        'goals': get_goals(family_id),
        'budgets': get_budgets(family_id),
        'wallet_report': get_wallet_report(family_id),
        'member_report': get_member_report(family_id),
        'currency_report': get_currency_report(family_id),
        'recent': get_recent_transactions(family_id),
        'audit_logs': get_audit_logs(family_id),
        'debt_payments': get_debt_payments(family_id),
        'goal_contributions': get_goal_contributions(family_id),
        'notification_settings': get_notification_settings(user),
        'scheduled_payments': get_scheduled_payments(family_id),
        'financial_plan': get_financial_plan_items(family_id),
        'financial_calendar': get_financial_calendar(family_id),
        'mandatory_payments': get_mandatory_payments_month(family_id),
        'scheduled_payment_issues': get_scheduled_payment_issues(family_id),
        'month_end_money': get_money_until_month_end(family_id),
        'budget_wizard': get_budget_wizard_profile(family_id),
        'ai_rules': get_ai_personal_rules(family_id) if has_permission(user, 'view_ai_analysis') else [],
        'ai_analysis': get_ai_monthly_analysis_with_rules(user) if has_permission(user, 'view_ai_analysis') else None,
    }


async def notify_admin_error(message: str):
    """Send critical error to Telegram admin. Best-effort: never breaks request handling."""
    if not ERROR_NOTIFY_TELEGRAM or not BOT_TOKEN or not ADMIN_ID:
        return
    try:
        safe = str(message)[:3500]
        async with ClientSession() as session:
            await session.post(
                f'https://api.telegram.org/bot{BOT_TOKEN}/sendMessage',
                json={'chat_id': ADMIN_ID, 'text': '🚨 FamilyFinance error\n\n' + safe},
                timeout=8,
            )
    except Exception:
        logger.exception('Could not notify admin about error')

async def notify_restore_otp(code: str, ip: str):
    """Send restore OTP only to Telegram superadmin/admin. The API never returns the code."""
    chat_id = SUPERADMIN_TELEGRAM_ID or ADMIN_ID
    if not BOT_TOKEN or not chat_id:
        raise ValueError('SUPERADMIN_TELEGRAM_ID или ADMIN_ID не настроен: некуда отправить restore-код')
    text = (
        '🔐 FamilyFinance restore OTP\n\n'
        f'Код: {code}\n'
        f'IP: {ip}\n'
        f'Действует: {RESTORE_OTP_TTL_SECONDS} сек.\n\n'
        'Не передавайте код никому.'
    )
    async with ClientSession() as session:
        resp = await session.post(
            f'https://api.telegram.org/bot{BOT_TOKEN}/sendMessage',
            json={'chat_id': chat_id, 'text': text},
            timeout=8,
        )
        if resp.status >= 400:
            raise ValueError('Telegram не принял restore OTP. Проверьте SUPERADMIN_TELEGRAM_ID/ADMIN_ID')


async def notify_budget_family(family_id: int, text: str):
    """Best-effort Telegram уведомление только тем участникам, у кого включены бюджетные уведомления."""
    if not BOT_TOKEN:
        return
    try:
        members = get_budget_notification_recipients(family_id)
        async with ClientSession() as session:
            for m in members:
                tg_id = m.get('telegram_id')
                if tg_id:
                    await session.post(
                        f'https://api.telegram.org/bot{BOT_TOKEN}/sendMessage',
                        json={'chat_id': tg_id, 'text': text[:3500]},
                        timeout=8,
                    )
    except Exception:
        logger.exception('Could not send budget notifications')

async def notify_budget_if_needed(family_id: int):
    alerts = get_budget_alerts_detailed(family_id)
    messages = []
    for a in alerts:
        if should_send_budget_alert(family_id, int(a['budget_id']), float(a['percent']), BUDGET_NOTIFY_COOLDOWN_HOURS):
            messages.append(f"⚠️ Превышен бюджет: {a['category_name']} — {a['percent']}% ({a['spent_base']} из {a['limit_base']})")
    if messages:
        await notify_budget_family(family_id, 'Бюджет семьи:\n' + '\n'.join(messages[:5]))

async def api_wallets(request):
    try:
        user = auth_user(request)
        if request.method == 'GET':
            return web.json_response({'ok': True, 'wallets': get_wallets(fid(user)), 'wallet_report': get_wallet_report(fid(user))})
        data = await request.json()
        wallet_id = add_wallet(user, data.get('name',''), data.get('currency','UZS'), data.get('initial_balance') or 0, data.get('include_in_free_money', True))
        return web.json_response({'ok': True, 'wallet_id': wallet_id, **init_payload(get_or_create_user(user['telegram_id'], user.get('full_name','Пользователь')) )})
    except ValueError as e:
        return json_error(str(e), 400)

async def api_wallet_update(request):
    try:
        user = auth_user(request)
        data = await request.json()
        update_wallet(user, int(request.match_info['wallet_id']), data.get('name',''), data.get('include_in_free_money') if 'include_in_free_money' in data else None)
        return web.json_response({'ok': True})
    except ValueError as e:
        return json_error(str(e), 400)

async def api_wallet_delete(request):
    try:
        user = auth_user(request)
        delete_wallet(user, int(request.match_info['wallet_id']))
        return web.json_response({'ok': True})
    except ValueError as e:
        return json_error(str(e), 400)

async def api_categories(request):
    try:
        user = auth_user(request)
        if request.method == 'GET':
            return web.json_response({'ok': True, 'income_categories': get_categories(fid(user), 'income'), 'expense_categories': get_categories(fid(user), 'expense')})
        data = await request.json()
        category_id = add_category(user, data.get('name',''), data.get('type','expense'), data.get('parent_id'))
        return web.json_response({'ok': True, 'category_id': category_id})
    except ValueError as e:
        return json_error(str(e), 400)

async def api_category_update(request):
    try:
        user = auth_user(request)
        data = await request.json()
        update_category(user, int(request.match_info['category_id']), data.get('name',''))
        return web.json_response({'ok': True})
    except ValueError as e:
        return json_error(str(e), 400)

async def api_category_delete(request):
    try:
        user = auth_user(request)
        delete_category(user, int(request.match_info['category_id']))
        return web.json_response({'ok': True})
    except ValueError as e:
        return json_error(str(e), 400)

async def api_operations(request):
    try:
        user = auth_user(request)
        items = get_transactions_filtered(
            fid(user),
            q=request.query.get('q',''),
            tx_type=request.query.get('type',''),
            wallet_id=request.query.get('wallet_id') or None,
            category_id=request.query.get('category_id') or None,
            user_id=request.query.get('user_id') or None,
            currency=request.query.get('currency',''),
            date_from=request.query.get('date_from'),
            date_to=request.query.get('date_to'),
            limit=request.query.get('limit') or 100,
        )
        return web.json_response({'ok': True, 'operations': items})
    except ValueError as e:
        return json_error(str(e), 400)

async def api_admin_backup_download(request):
    require_admin_panel(request)
    path = get_backup_path(request.match_info['name'])
    return web.FileResponse(path, headers={'Content-Disposition': f'attachment; filename="{path.name}"'})


async def api_admin_restore_request(request):
    global _restore_maintenance_until
    require_superadmin(request)
    code = f'{secrets.randbelow(1000000):06d}'
    _restore_otps[code] = {'expires': time() + RESTORE_OTP_TTL_SECONDS, 'ip': _client_ip(request)}
    _restore_maintenance_until = time() + RESTORE_OTP_TTL_SECONDS
    try:
        await notify_restore_otp(code, _client_ip(request))
    except Exception as e:
        admin_log('restore_otp_failed', admin_label='superadmin', ip_address=_client_ip(request), details=str(e))
        raise
    admin_log('restore_otp_request', admin_label='superadmin', ip_address=_client_ip(request), details='OTP sent to Telegram; maintenance enabled')
    return web.json_response({'ok': True, 'message': 'Restore-код отправлен superadmin в Telegram', 'expires_in': RESTORE_OTP_TTL_SECONDS, 'maintenance_mode': True})

def _consume_restore_otp(code: str, request: web.Request):
    data = _restore_otps.pop(str(code or '').strip(), None)
    if not data:
        raise ValueError('Неверный или уже использованный restore-код')
    if time() > data['expires']:
        raise ValueError('Restore-код истек')
    if data.get('ip') != _client_ip(request):
        raise ValueError('Restore-код создан с другого IP')

async def api_admin_backup_restore(request):
    global _restore_maintenance_until
    require_superadmin(request)
    if time() > _restore_maintenance_until:
        admin_log('restore_failed', admin_label='superadmin', ip_address=_client_ip(request), details='maintenance mode not active')
        return json_error('Restore запрещен: maintenance mode не активен. Сначала запросите OTP.', 409)
    try:
        reader = await request.multipart()
        otp_field = await reader.next()
        if not otp_field or otp_field.name != 'otp':
            admin_log('restore_failed', admin_label='superadmin', ip_address=_client_ip(request), details='missing otp')
            return json_error('Нужен одноразовый restore-код', 400)
        otp = (await otp_field.text()).strip()
        _consume_restore_otp(otp, request)
        admin_log('restore_confirm', admin_label='superadmin', ip_address=_client_ip(request), details='OTP accepted')
        field = await reader.next()
        if not field or field.name != 'backup':
            admin_log('restore_failed', admin_label='superadmin', ip_address=_client_ip(request), details='missing file')
            return json_error('Нужно загрузить файл backup', 400)
        suffix = Path(field.filename or 'backup').suffix or '.backup'
        fd, tmp_name = tempfile.mkstemp(prefix='restore_', suffix=suffix)
        os.close(fd)
        try:
            max_size = 25 * 1024 * 1024
            total = 0
            with open(tmp_name, 'wb') as f:
                while True:
                    chunk = await field.read_chunk()
                    if not chunk:
                        break
                    total += len(chunk)
                    if total > max_size:
                        raise ValueError('Backup-файл слишком большой')
                    f.write(chunk)
            pre_restore_backup = create_backup()
            admin_log('pre_restore_backup', admin_label='superadmin', ip_address=_client_ip(request), details=str(pre_restore_backup))
            msg = restore_backup_from_file(tmp_name)
            _restore_maintenance_until = 0
            full_msg = f'{msg}. Pre-restore backup: {pre_restore_backup}'
            logger.warning('Database restore completed by superadmin: %s', full_msg)
            admin_log('restore_backup', admin_label='superadmin', ip_address=_client_ip(request), details=full_msg)
            return web.json_response({'ok': True, 'message': full_msg, 'pre_restore_backup': pre_restore_backup, 'backups': list_backups(), 'maintenance_mode': False})
        finally:
            try:
                os.remove(tmp_name)
            except OSError:
                pass
    except ValueError as e:
        admin_log('restore_failed', admin_label='superadmin', ip_address=_client_ip(request), details=str(e))
        return json_error(str(e), 400)

async def api_admin_user_role(request):
    require_admin_panel(request)
    data = await request.json()
    admin_set_user_role(int(data.get('user_id') or 0), data.get('role','member'))
    admin_log('set_user_role', admin_label='admin', ip_address=_client_ip(request), entity_type='user', entity_id=int(data.get('user_id') or 0), details=data.get('role','member'))
    return web.json_response({'ok': True, 'users': get_admin_users()})


async def api_delete_account(request):
    try:
        user = auth_user(request)
        data = await request.json()
        result = delete_my_account(user, data.get('confirm', ''))
        return web.json_response({'ok': True, **result})
    except ValueError as e:
        return json_error(str(e), 400)

async def api_delete_family(request):
    try:
        user = auth_user(request)
        data = await request.json()
        result = delete_my_family(user, data.get('confirm', ''))
        return web.json_response({'ok': True, **result})
    except ValueError as e:
        return json_error(str(e), 400)


async def api_init(request):
    try:
        return web.json_response(init_payload(auth_user(request)))
    except ValueError as e:
        return json_error(str(e), 401)

async def api_reports(request):
    try:
        user = auth_user(request)
        month = request.query.get('month')
        date_from = request.query.get('date_from')
        date_to = request.query.get('date_to')
        family_id = fid(user)
        # Для старых месячных графиков оставляем month. Для расширенных отчетов поддерживаем диапазон дат.
        return web.json_response({
            'ok': True,
            'month_summary': get_period_summary(family_id, month, date_from, date_to),
            'category_report': get_expense_by_categories(family_id, month, date_from, date_to),
            'daily_chart': get_daily_chart(family_id, month, date_from, date_to),
            'chart_pack': get_webapp_chart_pack(family_id, month, date_from, date_to),
            'forecast': get_expense_forecast(family_id, month),
            'budgets': get_budgets(family_id, month),
            'wallet_report': get_wallet_report(family_id),
            'member_report': get_member_report(family_id, month, date_from, date_to),
            'currency_report': get_currency_report(family_id, month, date_from, date_to),
            'debts': get_debts(family_id),
            'goals': get_goals(family_id),
            'debt_payments': get_debt_payments(family_id),
            'goal_contributions': get_goal_contributions(family_id)
        })
    except ValueError as e:
        return json_error(str(e), 400)

async def api_transaction(request):
    try:
        user = auth_user(request)
        data = await request.json()
        tx_id = add_transaction(user, data.get('type'), data.get('amount'), data.get('currency','UZS'), int(data.get('wallet_id') or 0), int(data.get('category_id') or 0), data.get('comment',''))
        await notify_budget_if_needed(fid(user))
        return web.json_response({'ok': True, 'transaction_id': tx_id})
    except ValueError as e:
        return json_error(str(e), 400)
    except Exception as e:
        return json_error(f'Ошибка сохранения: {e}', 400)

async def api_transaction_edit(request):
    try:
        user = auth_user(request)
        data = await request.json()
        edit_transaction(user, int(request.match_info['tx_id']), data.get('amount'), int(data.get('wallet_id') or 0), int(data.get('category_id') or 0), data.get('comment',''))
        return web.json_response({'ok': True})
    except ValueError as e:
        return json_error(str(e), 400)

async def api_transaction_delete(request):
    try:
        user = auth_user(request)
        delete_transaction(user, int(request.match_info['tx_id']))
        return web.json_response({'ok': True})
    except ValueError as e:
        return json_error(str(e), 400)

async def api_transfer_delete(request):
    try:
        user = auth_user(request)
        delete_transfer(user, int(request.match_info['transfer_id']))
        return web.json_response({'ok': True})
    except ValueError as e:
        return json_error(str(e), 400)

async def api_transfer(request):
    try:
        user = auth_user(request)
        data = await request.json()
        amount_to = data.get('amount_to')
        amount_to = None if amount_to in (None, '') else float(amount_to)
        transfer_between_wallets(user, int(data.get('from_wallet_id') or 0), int(data.get('to_wallet_id') or 0), data.get('amount_from'), amount_to, data.get('comment',''))
        return web.json_response({'ok': True})
    except ValueError as e:
        return json_error(str(e), 400)

async def api_join_family(request):
    try:
        user = auth_user(request)
        data = await request.json()
        join_family(user, data.get('invite_code',''), data.get('role','wife'))
        return web.json_response({'ok': True})
    except ValueError as e:
        return json_error(str(e), 400)

async def api_family_member_role(request):
    try:
        user = auth_user(request)
        data = await request.json()
        members = update_family_member_role(user, int(data.get('member_id') or 0), data.get('role','member'))
        return web.json_response({'ok': True, 'members': members})
    except ValueError as e:
        return json_error(str(e), 400)


async def api_family_member_remove(request):
    try:
        user = auth_user(request)
        data = await request.json()
        members = remove_family_member(user, int(data.get('member_id') or 0))
        return web.json_response({'ok': True, 'members': members})
    except ValueError as e:
        return json_error(str(e), 400)


async def api_family_member_permissions(request):
    try:
        user = auth_user(request)
        member_id = int(request.match_info.get('member_id') or 0)
        if request.method == 'GET':
            return web.json_response({'ok': True, **get_member_permissions(user, member_id)})
        data = await request.json()
        return web.json_response({'ok': True, **set_member_permissions(user, member_id, data.get('permissions') or [])})
    except ValueError as e:
        return json_error(str(e), 400)


async def api_rates(request):
    try:
        user = auth_user(request)
        data = await request.json()
        set_rate(user, data.get('currency','UZS'), data.get('rate_to_base') or 1)
        return web.json_response({'ok': True})
    except ValueError as e:
        return json_error(str(e), 400)

async def api_debt(request):
    try:
        user = auth_user(request)
        data = await request.json()
        amount = data.get('amount', data.get('total_amount', 0))
        debt_id = add_debt(user, data.get('name',''), amount or 0, data.get('currency','USD'), data.get('comment',''))
        return web.json_response({'ok': True, 'debt_id': debt_id})
    except ValueError as e:
        return json_error(str(e), 400)

async def api_pay_debt(request):
    try:
        user = auth_user(request)
        data = await request.json()
        pay_debt(user, int(data.get('debt_id') or 0), data.get('amount') or 0, int(data.get('wallet_id') or 0))
        await notify_budget_if_needed(fid(user))
        return web.json_response({'ok': True})
    except ValueError as e:
        return json_error(str(e), 400)

async def api_goal(request):
    try:
        user = auth_user(request)
        data = await request.json()
        goal_id = add_goal(user, data.get('name',''), data.get('target_amount') or 0, data.get('currency','USD'), data.get('deadline',''))
        return web.json_response({'ok': True, 'goal_id': goal_id})
    except ValueError as e:
        return json_error(str(e), 400)

async def api_goal_add(request):
    try:
        user = auth_user(request)
        data = await request.json()
        add_goal_money(user, int(data.get('goal_id') or 0), data.get('amount') or 0, int(data.get('wallet_id') or 0) if data.get('wallet_id') else None)
        await notify_budget_if_needed(fid(user))
        return web.json_response({'ok': True})
    except ValueError as e:
        return json_error(str(e), 400)

async def api_budget(request):
    try:
        user = auth_user(request)
        data = await request.json()
        set_budget(user=user, category_id=int(data.get('category_id') or 0), month=data.get('month'), limit_amount=data.get('limit_amount') or 0, currency=data.get('currency','UZS'))
        return web.json_response({'ok': True})
    except ValueError as e:
        return json_error(str(e), 400)

async def api_export(request):
    try:
        user = auth_user(request)
        require_permission(user, 'export')
        month = request.query.get('month')
        date_from = request.query.get('date_from')
        date_to = request.query.get('date_to')
        family_id = fid(user)
        period = get_period_summary(family_id, month, date_from, date_to)
    except ValueError as e:
        return json_error(str(e), 401)

    def style_header(ws):
        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='1F2937')
            cell.alignment = Alignment(horizontal='center')

    def autosize(ws):
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = min(max(len(str(c.value or '')) for c in col) + 2, 40)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Операции'
    ws.append(['Дата','Тип','Сумма','Валюта','Сумма в базе','Кошелек','Категория','Кто добавил','Комментарий'])
    style_header(ws)
    for t in get_transactions_for_export(family_id, month, date_from, date_to):
        ws.append([str(t.get('created_at',''))[:19], t.get('type'), t.get('amount'), t.get('currency'), t.get('amount_base'), t.get('wallet_name'), t.get('category_name'), t.get('user_name'), t.get('comment')])
    autosize(ws)

    ws_wallets = wb.create_sheet('Кошельки')
    ws_wallets.append(['Кошелек', 'Валюта', 'Остаток', 'Эквивалент в базе'])
    style_header(ws_wallets)
    for w in get_wallet_report(family_id):
        ws_wallets.append([w['name'], w['currency'], w.get('balance', 0), w.get('balance_base', 0)])
    autosize(ws_wallets)

    ws_members = wb.create_sheet('Члены семьи')
    ws_members.append(['Имя', 'Роль', 'Доход', 'Расход', 'Итог', 'Кол-во операций'])
    style_header(ws_members)
    for m in get_member_report(family_id, month, date_from, date_to):
        ws_members.append([m['full_name'], m['role'], m['income'], m['expense'], m['balance'], m['count']])
    autosize(ws_members)

    ws_cur = wb.create_sheet('Валюты')
    ws_cur.append(['Валюта', 'Доход', 'Расход', 'Доход в базе', 'Расход в базе', 'Остаток в кошельках'])
    style_header(ws_cur)
    for c in get_currency_report(family_id, month, date_from, date_to):
        ws_cur.append([c['currency'], c['income'], c['expense'], c['income_base'], c['expense_base'], c['wallet_balance']])
    autosize(ws_cur)

    ws2 = wb.create_sheet('Итоги')
    ws2.append(['Показатель','Значение'])
    style_header(ws2)
    s = get_summary(family_id)
    ws2.append(['Период', period['period']])
    ws2.append(['Валюта отчета', s['base_currency']])
    ws2.append(['Доход за период', period['income']])
    ws2.append(['Расход за период', period['expense']])
    ws2.append(['Остаток за период', period['balance']])
    ws2.append(['Реальный остаток в кошельках', s['balance']])
    ws2.append(['Долги осталось', s['debt_left']])
    autosize(ws2)

    suffix = (month or 'period').replace('-', '_')
    bio = BytesIO(); wb.save(bio); bio.seek(0)
    return web.Response(body=bio.read(), headers={'Content-Type':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet','Content-Disposition':f'attachment; filename=family_finance_export_{suffix}.xlsx'})


@web.middleware
async def security_headers_middleware(request, handler):
    response = await handler(request)
    response.headers.setdefault('Content-Security-Policy', "default-src 'self'; script-src 'self' https://telegram.org; style-src 'self'; img-src 'self' data:; connect-src 'self'; font-src 'self'; frame-ancestors 'none'; base-uri 'self'; form-action 'self'")
    response.headers.setdefault('X-Frame-Options', 'DENY')
    response.headers.setdefault('X-Content-Type-Options', 'nosniff')
    response.headers.setdefault('Referrer-Policy', 'no-referrer')
    response.headers.setdefault('Permissions-Policy', 'camera=(), microphone=(), geolocation=()')
    return response

@web.middleware
async def error_middleware(request, handler):
    try:
        return await handler(request)
    except web.HTTPException:
        raise
    except ValueError as e:
        logger.warning('Validation error %s %s: %s', request.method, request.path, e)
        return json_error(str(e), 400)
    except Exception as exc:
        logger.exception('Unhandled error %s %s', request.method, request.path)
        _capture_exception(exc)
        await notify_admin_error(f'{request.method} {request.path} failed. Check logs/app.log')
        return json_error('Внутренняя ошибка сервера. Подробности записаны в лог.', 500)

def _new_admin_session(is_superadmin: bool = False, ip_address: str | None = None):
    session_id = secrets.token_urlsafe(32)
    csrf = secrets.token_urlsafe(24)
    admin_session_create(session_id, csrf, bool(is_superadmin), ip_address, ADMIN_SESSION_TTL_SECONDS)
    return session_id, csrf

def _get_admin_session(request: web.Request):
    sid = request.cookies.get('ff_admin_session') or ''
    sess = admin_session_get(sid)
    if not sess:
        return None
    # Session hijack guard: bind admin session to the login IP when available.
    if sess.get('ip') and sess.get('ip') != _client_ip(request):
        admin_session_delete(sid)
        return None
    return sess

def admin_authorized(request: web.Request) -> bool:
    # Preferred: session cookie. Legacy header token remains for CLI/smoke tests.
    if _get_admin_session(request):
        return True
    token = request.headers.get('X-Admin-Token')
    return bool(ADMIN_PANEL_TOKEN and ADMIN_PANEL_TOKEN != 'change-me' and token == ADMIN_PANEL_TOKEN)

def superadmin_authorized(request: web.Request) -> bool:
    sess = _get_admin_session(request)
    if sess and sess.get('is_superadmin'):
        return True
    token = request.headers.get('X-Superadmin-Token')
    return bool(SUPERADMIN_TOKEN and SUPERADMIN_TOKEN != 'change-me-superadmin' and token == SUPERADMIN_TOKEN)

def require_admin_panel(request: web.Request):
    if not admin_authorized(request):
        try:
            admin_log('admin_auth_failed', admin_label='unknown', ip_address=_client_ip(request), details=request.path)
        except Exception:
            pass
        raise web.HTTPUnauthorized(text='Unauthorized')
    # CSRF for browser session unsafe methods. Header token auth remains supported for scripts.
    sess = _get_admin_session(request)
    if sess and request.method not in ('GET', 'HEAD', 'OPTIONS'):
        sent = request.headers.get(ADMIN_CSRF_HEADER, '')
        if not hmac.compare_digest(sent, sess.get('csrf', '')):
            try:
                admin_log('csrf_failed', admin_label='admin', ip_address=_client_ip(request), details=request.path)
            except Exception:
                pass
            raise web.HTTPForbidden(text='CSRF failed')

def require_superadmin(request: web.Request):
    require_admin_panel(request)
    if not superadmin_authorized(request):
        try:
            admin_log('superadmin_auth_failed', admin_label='admin', ip_address=_client_ip(request), details=request.path)
        except Exception:
            pass
        raise web.HTTPForbidden(text='Superadmin required')

async def api_admin_login(request):
    data = await request.json()
    admin_token = str(data.get('admin_token') or '')
    super_token = str(data.get('superadmin_token') or '')
    if not (ADMIN_PANEL_TOKEN and ADMIN_PANEL_TOKEN != 'change-me' and hmac.compare_digest(admin_token, ADMIN_PANEL_TOKEN)):
        admin_log('admin_login_failed', admin_label='unknown', ip_address=_client_ip(request), details='bad admin token')
        raise web.HTTPUnauthorized(text='Unauthorized')
    is_super = bool(SUPERADMIN_TOKEN and SUPERADMIN_TOKEN != 'change-me-superadmin' and hmac.compare_digest(super_token, SUPERADMIN_TOKEN))
    sid, csrf = _new_admin_session(is_super, _client_ip(request))
    admin_log('admin_login', admin_label='superadmin' if is_super else 'admin', ip_address=_client_ip(request), details='session created')
    resp = web.json_response({'ok': True, 'csrf': csrf, 'is_superadmin': is_super, 'ttl': ADMIN_SESSION_TTL_SECONDS})
    resp.set_cookie('ff_admin_session', sid, max_age=ADMIN_SESSION_TTL_SECONDS, httponly=True, secure=SESSION_COOKIE_SECURE, samesite='Strict', path='/')
    return resp

async def api_admin_logout(request):
    sess_id = request.cookies.get('ff_admin_session') or ''
    if sess_id:
        admin_session_delete(sess_id)
    admin_log('admin_logout', admin_label='admin', ip_address=_client_ip(request), details='session closed')
    resp = web.json_response({'ok': True})
    resp.del_cookie('ff_admin_session', path='/')
    return resp

async def admin_index(request):
    return web.FileResponse(ADMIN_DIR / 'admin.html')


async def api_admin_backup(request):
    require_admin_panel(request)
    path = create_backup()
    deleted = cleanup_old_backups()
    logger.info('Manual backup created: %s, old deleted: %s', path, deleted)
    admin_log('create_backup', admin_label='admin', ip_address=_client_ip(request), details=str(path))
    return web.json_response({'ok': True, 'path': path, 'deleted_old': deleted, 'backups': list_backups()})


async def api_operation_history(request):
    try:
        user = auth_user(request)
        return web.json_response({'ok': True, 'history': get_operation_history(user, int(request.match_info['tx_id']))})
    except ValueError as e:
        return json_error(str(e), 400)

async def api_audit_logs(request):
    try:
        user = auth_user(request)
        logs = get_audit_logs_filtered(
            fid(user),
            action=request.query.get('action',''),
            user_id=request.query.get('user_id') or None,
            entity_type=request.query.get('entity_type',''),
            entity_id=request.query.get('entity_id') or None,
            date_from=request.query.get('date_from'),
            date_to=request.query.get('date_to'),
            limit=request.query.get('limit') or 200,
        )
        return web.json_response({'ok': True, 'audit_logs': logs})
    except ValueError as e:
        return json_error(str(e), 400)

async def api_undo(request):
    try:
        user = auth_user(request)
        data = await request.json()
        undo_audit_action(user, int(data.get('audit_id') or 0))
        return web.json_response({'ok': True, **init_payload(get_or_create_user(user['telegram_id'], user.get('full_name','Пользователь')))})
    except ValueError as e:
        message = str(e)
        if 'недостаточно средств' in message.lower() or 'нельзя сделать undo' in message.lower():
            return web.json_response({
                'ok': False,
                'error': message,
                'suggestion': 'Полная отмена невозможна без отрицательного баланса. Создайте ручную корректирующую операцию на доступную сумму или сначала пополните нужный кошелек.',
                'correction_available': True,
            }, status=409)
        return json_error(message, 400)


# --- Level 4.1 API: admin charts, family detail, user blocking, audit export ---
_orig_auth_user_l41 = auth_user

async def api_admin_user_block(request):
    require_admin_panel(request)
    data = await request.json()
    user_id = int(data.get('user_id') or 0)
    blocked = bool(data.get('blocked'))
    admin_set_user_blocked(user_id, blocked)
    admin_log('block_user' if blocked else 'unblock_user', admin_label='admin', ip_address=_client_ip(request), entity_type='user', entity_id=user_id, details=str(blocked))
    return web.json_response({'ok': True, 'users': get_admin_users()})

async def api_admin_family_operations(request):
    require_admin_panel(request)
    family_id = int(request.match_info['family_id'])
    rows = get_admin_family_operations(
        family_id,
        q=request.query.get('q',''),
        tx_type=request.query.get('type',''),
        user_id=request.query.get('user_id') or None,
        date_from=request.query.get('date_from'),
        date_to=request.query.get('date_to'),
        limit=request.query.get('limit') or 200,
    )
    admin_log('view_family_operations', admin_label='admin', ip_address=_client_ip(request), entity_type='family', entity_id=family_id, details=f'count={len(rows)}')
    return web.json_response({'ok': True, 'operations': rows})

async def api_admin_audit_export(request):
    require_admin_panel(request)
    logs = get_admin_audit_logs_filtered(
        action=request.query.get('action',''),
        admin_label=request.query.get('admin_label',''),
        entity_type=request.query.get('entity_type',''),
        date_from=request.query.get('date_from'),
        date_to=request.query.get('date_to'),
        limit=request.query.get('limit') or 5000,
    )
    wb = Workbook()
    ws = wb.active
    ws.title = 'Admin audit logs'
    headers = ['ID','Admin','IP','Action','Entity type','Entity ID','Details','Created at']
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill('solid', fgColor='E8F1FF')
    for r in logs:
        ws.append([r.get('id'), r.get('admin_label'), r.get('ip_address'), r.get('action'), r.get('entity_type'), r.get('entity_id'), r.get('details'), r.get('created_at')])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = min(max(len(str(cell.value or '')) for cell in col) + 2, 60)
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    admin_log('export_admin_audit_logs', admin_label='admin', ip_address=_client_ip(request), details=f'rows={len(logs)}')
    return web.Response(
        body=bio.read(),
        headers={'Content-Disposition': 'attachment; filename="admin_audit_logs.xlsx"'},
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

# Override stats endpoint with Level 4.1 chart data.
async def api_admin_stats(request):
    require_admin_panel(request)
    admin_log('view_dashboard', admin_label='admin', ip_address=_client_ip(request), details='stats level4.1')
    return web.json_response({
        'ok': True,
        'stats': get_admin_stats(),
        'charts': get_admin_chart_data(30),
        'families': get_admin_families(),
        'users': get_admin_users(),
        'backups': list_backups(),
        'admin_audit_logs': get_admin_audit_logs(100),
    })

async def api_admin_family_detail(request):
    require_admin_panel(request)
    family_id = int(request.match_info['family_id'])
    admin_log('view_family_detail', admin_label='admin', ip_address=_client_ip(request), entity_type='family', entity_id=family_id, details='detail card')
    return web.json_response({'ok': True, **get_admin_family_detail(family_id)})

# --- Level 4.3 API: chart pack, forecast and PDF monthly family report ---
async def api_chart_pack(request):
    try:
        user = auth_user(request)
        month = request.query.get('month')
        date_from = request.query.get('date_from')
        date_to = request.query.get('date_to')
        return web.json_response({'ok': True, 'charts': get_webapp_chart_pack(fid(user), month, date_from, date_to)})
    except ValueError as e:
        return json_error(str(e), 400)

async def api_forecast(request):
    try:
        user = auth_user(request)
        month = request.query.get('month') or date.today().strftime('%Y-%m')
        return web.json_response({'ok': True, 'forecast': get_expense_forecast(fid(user), month)})
    except ValueError as e:
        return json_error(str(e), 400)

def _pdf_money(v):
    try:
        return f"{float(v or 0):,.2f}".replace(',', ' ')
    except Exception:
        return str(v or 0)

async def api_notification_settings(request):
    try:
        user = auth_user(request)
        if request.method == 'GET':
            return web.json_response({'ok': True, 'notification_settings': get_notification_settings(user)})
        data = await request.json()
        settings = update_notification_settings(
            user,
            daily_enabled=data.get('daily_enabled') if 'daily_enabled' in data else None,
            budget_alert_enabled=data.get('budget_alert_enabled') if 'budget_alert_enabled' in data else None,
            scheduled_payment_enabled=data.get('scheduled_payment_enabled') if 'scheduled_payment_enabled' in data else None,
        )
        return web.json_response({'ok': True, 'notification_settings': settings, **init_payload(get_or_create_user(user['telegram_id'], user.get('full_name','Пользователь')))})
    except ValueError as e:
        return json_error(str(e), 400)

async def api_monthly_pdf(request):
    try:
        user = auth_user(request)
        family_id = fid(user)
        month = request.query.get('month') or date.today().strftime('%Y-%m')
        family = get_family(family_id)
        summary = get_month_summary(family_id, month)
        categories = get_expense_by_categories(family_id, month)
        wallets = get_wallet_report(family_id)
        members = get_member_report(family_id, month)
        currencies = get_currency_report(family_id, month)
        debts = get_debts(family_id)
        goals_list = get_goals(family_id)
        budgets = get_budgets(family_id, month)
        forecast = get_expense_forecast(family_id, month)

        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.graphics.shapes import Drawing, Rect, String

        font_name = PDF_REPORT_FONT or 'Helvetica'
        bold_font_name = font_name + '-Bold'
        try:
            if PDF_REPORT_FONT_PATH and os.path.exists(PDF_REPORT_FONT_PATH):
                pdfmetrics.registerFont(TTFont(font_name, PDF_REPORT_FONT_PATH))
            if PDF_REPORT_FONT_BOLD_PATH and os.path.exists(PDF_REPORT_FONT_BOLD_PATH):
                pdfmetrics.registerFont(TTFont(bold_font_name, PDF_REPORT_FONT_BOLD_PATH))
            else:
                bold_font_name = font_name
        except Exception:
            logger.warning('Could not register PDF font, falling back to Helvetica', exc_info=True)
            font_name = 'Helvetica'
            bold_font_name = 'Helvetica-Bold'

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=14*mm, leftMargin=14*mm, topMargin=14*mm, bottomMargin=14*mm)
        styles = getSampleStyleSheet()
        for st in ['Title','Heading2','Normal']:
            styles[st].fontName = font_name
        styles.add(ParagraphStyle(name='Small', parent=styles['Normal'], fontName=font_name, fontSize=8, leading=10))
        story = []
        story.append(Paragraph(f"Family Finance — месячный отчет", styles['Title']))
        story.append(Paragraph(f"Семья: {family.get('name')} | Месяц: {month} | Базовая валюта: {family.get('base_currency')}", styles['Normal']))
        story.append(Spacer(1, 8))
        def add_table(title, headers, rows):
            story.append(Paragraph(title, styles['Heading2']))
            data = [headers] + rows
            tbl = Table(data, repeatRows=1, hAlign='LEFT')
            tbl.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#E8EEF7')),
                ('TEXTCOLOR', (0,0), (-1,0), colors.HexColor('#182230')),
                ('GRID', (0,0), (-1,-1), 0.25, colors.HexColor('#D0D5DD')),
                ('FONTNAME', (0,0), (-1,0), bold_font_name),
                ('FONTNAME', (0,1), (-1,-1), font_name),
                ('FONTSIZE', (0,0), (-1,-1), 8),
                ('VALIGN', (0,0), (-1,-1), 'TOP'),
                ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F9FAFB')]),
            ]))
            story.append(tbl)
            story.append(Spacer(1, 8))

        def add_bar_chart(title, labels, values, max_items=8):
            pairs = [(str(l or ''), float(v or 0)) for l, v in zip(labels, values) if float(v or 0) > 0][:max_items]
            if not pairs:
                return
            story.append(Paragraph(title, styles['Heading2']))
            width, row_h = 170*mm, 10*mm
            height = max(28*mm, (len(pairs)+1)*row_h)
            drawing = Drawing(width, height)
            max_v = max(v for _, v in pairs) or 1
            palette = ['#2563EB','#16A34A','#F97316','#DC2626','#7C3AED','#0891B2','#65A30D','#BE123C']
            y = height - row_h
            for idx, (label, value) in enumerate(pairs):
                bar_w = (width - 62*mm) * (value / max_v)
                drawing.add(String(0, y+2, label[:24], fontName=font_name, fontSize=8, fillColor=colors.HexColor('#344054')))
                drawing.add(Rect(45*mm, y, bar_w, 5*mm, fillColor=colors.HexColor(palette[idx % len(palette)]), strokeColor=None))
                drawing.add(String(48*mm + bar_w, y+1, _pdf_money(value), fontName=font_name, fontSize=8, fillColor=colors.HexColor('#475467')))
                y -= row_h
            story.append(drawing)
            story.append(Spacer(1, 8))

        add_table('Итоги', ['Доход', 'Расход', 'Баланс', 'Погашено долгов', 'Прогноз расходов'], [[_pdf_money(summary.get('income')), _pdf_money(summary.get('expense')), _pdf_money(summary.get('balance')), _pdf_money(summary.get('debt_paid')), _pdf_money(forecast.get('projected_expense'))]])
        add_bar_chart('График: расходы по категориям', [c.get('category_name') for c in categories], [c.get('amount') for c in categories])
        add_bar_chart('График: использование бюджетов', [b.get('category_name') for b in budgets], [b.get('percent') for b in budgets])
        add_table('Расходы по категориям', ['Категория', 'Сумма', '%'], [[c.get('category_name'), _pdf_money(c.get('amount')), c.get('percent')] for c in categories] or [['No data','','']])
        add_table('Кошельки', ['Кошелек', 'Баланс', 'Валюта', 'В базе'], [[w.get('name'), _pdf_money(w.get('balance')), w.get('currency'), _pdf_money(w.get('balance_base'))] for w in wallets] or [['No data','','','']])
        add_table('Члены семьи', ['Участник', 'Роль', 'Доход', 'Расход', 'Операции'], [[m.get('full_name'), m.get('role'), _pdf_money(m.get('income')), _pdf_money(m.get('expense')), m.get('count')] for m in members] or [['No data','','','','']])
        add_table('Валюты', ['Валюта', 'Доход', 'Расход', 'Баланс кошельков'], [[c.get('currency'), _pdf_money(c.get('income')), _pdf_money(c.get('expense')), _pdf_money(c.get('wallet_balance'))] for c in currencies] or [['No data','','','']])
        add_table('Бюджеты', ['Категория', 'Потрачено', 'Лимит', '%'], [[b.get('category_name'), _pdf_money(b.get('spent_base')), _pdf_money(b.get('limit_base')), b.get('percent')] for b in budgets] or [['No budgets','','','']])
        add_table('Долги', ['Название', 'Всего', 'Оплачено', 'Осталось'], [[d.get('name'), _pdf_money(d.get('total_amount')), _pdf_money(d.get('paid_amount')), _pdf_money(d.get('left_amount'))] for d in debts] or [['No debts','','','']])
        add_table('Цели', ['Название', 'Сейчас', 'Цель', 'Валюта'], [[g.get('name'), _pdf_money(g.get('current_amount')), _pdf_money(g.get('target_amount')), g.get('currency')] for g in goals_list] or [['No goals','','','']])
        story.append(Paragraph(f"Generated at {datetime.now().isoformat(timespec='seconds')}", styles['Small']))
        doc.build(story)
        pdf = buffer.getvalue()
        return web.Response(body=pdf, content_type='application/pdf', headers={'Content-Disposition': f'attachment; filename="family_report_{month}.pdf"'})
    except ValueError as e:
        return json_error(str(e), 400)
    except Exception as e:
        logger.exception('PDF report failed')
        return json_error('Не удалось сформировать PDF отчет', 500)


async def api_scheduled_payments(request):
    try:
        user = auth_user(request)
        if request.method == 'GET':
            return web.json_response({'ok': True, 'scheduled_payments': get_scheduled_payments(fid(user))})
        data = await request.json()
        sid = add_scheduled_payment(user, data.get('title',''), data.get('amount') or 0, data.get('currency','UZS'), data.get('due_day') or 1, data.get('kind','expense'), data.get('wallet_id'), data.get('category_id'), data.get('auto_create_expense'))
        return web.json_response({'ok': True, 'schedule_id': sid, **init_payload(get_or_create_user(user['telegram_id'], user.get('full_name','Пользователь')))})
    except ValueError as e:
        return json_error(str(e), 400)

async def api_scheduled_payment_update(request):
    try:
        user = auth_user(request)
        data = await request.json()
        update_scheduled_payment(user, int(request.match_info['schedule_id']), **data)
        return web.json_response({'ok': True, **init_payload(get_or_create_user(user['telegram_id'], user.get('full_name','Пользователь')))})
    except ValueError as e:
        return json_error(str(e), 400)

async def api_scheduled_payment_delete(request):
    try:
        user = auth_user(request)
        delete_scheduled_payment(user, int(request.match_info['schedule_id']))
        return web.json_response({'ok': True, **init_payload(get_or_create_user(user['telegram_id'], user.get('full_name','Пользователь')))})
    except ValueError as e:
        return json_error(str(e), 400)

async def api_financial_plan(request):
    try:
        user = auth_user(request)
        if request.method == 'GET':
            return web.json_response({'ok': True, 'financial_plan': get_financial_plan_items(fid(user))})
        data = await request.json()
        item_id = add_financial_plan_item(
            user,
            data.get('title',''),
            data.get('target_amount') or 0,
            data.get('currency','UZS'),
            data.get('current_amount') or 0,
            data.get('priority') or 3,
            data.get('deadline',''),
            data.get('note',''),
        )
        return web.json_response({'ok': True, 'plan_item_id': item_id, **init_payload(get_or_create_user(user['telegram_id'], user.get('full_name','Пользователь')))})
    except ValueError as e:
        return json_error(str(e), 400)

async def api_financial_plan_update(request):
    try:
        user = auth_user(request)
        data = await request.json()
        update_financial_plan_item(user, int(request.match_info['item_id']), **data)
        return web.json_response({'ok': True, **init_payload(get_or_create_user(user['telegram_id'], user.get('full_name','Пользователь')))})
    except ValueError as e:
        return json_error(str(e), 400)

async def api_financial_plan_delete(request):
    try:
        user = auth_user(request)
        delete_financial_plan_item(user, int(request.match_info['item_id']))
        return web.json_response({'ok': True, **init_payload(get_or_create_user(user['telegram_id'], user.get('full_name','Пользователь')))})
    except ValueError as e:
        return json_error(str(e), 400)



async def api_financial_calendar(request):
    try:
        user = auth_user(request)
        return web.json_response({'ok': True, 'calendar': get_financial_calendar(fid(user), request.query.get('month'))})
    except ValueError as e:
        return json_error(str(e), 400)

async def api_ai_analysis(request):
    try:
        user = auth_user(request)
        require_permission(user, 'view_ai_analysis')
        return web.json_response({'ok': True, 'analysis': get_ai_monthly_analysis_with_rules(user, request.query.get('month'))})
    except ValueError as e:
        return json_error(str(e), 400)


async def api_mandatory_payments(request):
    try:
        user = auth_user(request)
        require_permission(user, 'manage_schedules')
        return web.json_response({'ok': True, 'mandatory_payments': get_mandatory_payments_month(fid(user), request.query.get('month'))})
    except ValueError as e:
        return json_error(str(e), 400)


async def api_scheduled_payment_issues(request):
    try:
        user = auth_user(request)
        return web.json_response({'ok': True, 'issues': get_scheduled_payment_issues(fid(user), request.query.get('limit') or 50)})
    except Exception as exc:
        _capture_exception(exc)
        return web.json_response({'ok': False, 'error': str(exc)}, status=400)

async def api_month_end_money(request):
    try:
        user = auth_user(request)
        return web.json_response({'ok': True, 'month_end_money': get_money_until_month_end(fid(user), request.query.get('month'))})
    except Exception as exc:
        _capture_exception(exc)
        return web.json_response({'ok': False, 'error': str(exc)}, status=400)

async def api_mandatory_payment_pay(request):
    try:
        user = auth_user(request)
        data = await request.json()
        tx_id = pay_mandatory_payment(user, int(request.match_info['schedule_id']), data.get('wallet_id'), data.get('category_id'))
        fresh = get_or_create_user(user['telegram_id'], user.get('full_name','Пользователь'))
        return web.json_response({'ok': True, 'transaction_id': tx_id, **init_payload(fresh)})
    except Exception as exc:
        _capture_exception(exc)
        return web.json_response({'ok': False, 'error': str(exc)}, status=400)

async def api_mandatory_payment_mark_paid(request):
    try:
        user = auth_user(request)
        data = await request.json()
        tx_id = mark_mandatory_payment_paid(user, int(request.match_info['schedule_id']), data.get('wallet_id'), data.get('category_id'), data.get('transaction_id'), bool(data.get('zero_note')))
        fresh = get_or_create_user(user['telegram_id'], user.get('full_name','Пользователь'))
        return web.json_response({'ok': True, 'transaction_id': tx_id, **init_payload(fresh)})
    except Exception as exc:
        _capture_exception(exc)
        return web.json_response({'ok': False, 'error': str(exc)}, status=400)


async def api_mandatory_payment_linkable_transactions(request):
    try:
        user = auth_user(request)
        rows = get_linkable_transactions_for_mandatory(user, int(request.match_info['schedule_id']), request.query.get('month'))
        return web.json_response({'ok': True, 'transactions': rows})
    except Exception as exc:
        _capture_exception(exc)
        return web.json_response({'ok': False, 'error': str(exc)}, status=400)

async def api_mandatory_payment_link_existing(request):
    try:
        user = auth_user(request)
        data = await request.json()
        tx_id = link_existing_transaction_to_mandatory(user, int(request.match_info['schedule_id']), int(data.get('transaction_id') or 0), data.get('month'))
        fresh = get_or_create_user(user['telegram_id'], user.get('full_name','Пользователь'))
        return web.json_response({'ok': True, 'transaction_id': tx_id, **init_payload(fresh)})
    except Exception as exc:
        _capture_exception(exc)
        return web.json_response({'ok': False, 'error': str(exc)}, status=400)

async def api_scheduled_payment_issue_resolve(request):
    try:
        user = auth_user(request)
        resolve_scheduled_payment_issue(user, int(request.match_info['issue_id']))
        fresh = get_or_create_user(user['telegram_id'], user.get('full_name','Пользователь'))
        return web.json_response({'ok': True, **init_payload(fresh)})
    except Exception as exc:
        _capture_exception(exc)
        return web.json_response({'ok': False, 'error': str(exc)}, status=400)

async def api_scheduled_payment_retry(request):
    try:
        user = auth_user(request)
        tx_id = retry_scheduled_payment(user, int(request.match_info['schedule_id']))
        fresh = get_or_create_user(user['telegram_id'], user.get('full_name','Пользователь'))
        return web.json_response({'ok': True, 'transaction_id': tx_id, **init_payload(fresh)})
    except Exception as exc:
        _capture_exception(exc)
        return web.json_response({'ok': False, 'error': str(exc)}, status=400)

async def api_scheduled_payment_disable_auto(request):
    try:
        user = auth_user(request)
        disable_scheduled_payment_auto_create(user, int(request.match_info['schedule_id']))
        fresh = get_or_create_user(user['telegram_id'], user.get('full_name','Пользователь'))
        return web.json_response({'ok': True, **init_payload(fresh)})
    except Exception as exc:
        _capture_exception(exc)
        return web.json_response({'ok': False, 'error': str(exc)}, status=400)

async def api_ai_rules(request):
    try:
        user = auth_user(request)
        require_permission(user, 'manage_ai_rules')
        if request.method == 'GET':
            return web.json_response({'ok': True, 'ai_rules': get_ai_personal_rules(fid(user))})
        data = await request.json()
        rid = add_ai_personal_rule(user, data.get('title',''), data.get('rule_type','category_limit'), data.get('category_id'), data.get('threshold_amount') or 0, data.get('currency','UZS'), data.get('enabled', True))
        return web.json_response({'ok': True, 'rule_id': rid, **init_payload(get_or_create_user(user['telegram_id'], user.get('full_name','Пользователь')))})
    except ValueError as e:
        return json_error(str(e), 400)

async def api_ai_rule_update(request):
    try:
        user = auth_user(request)
        data = await request.json()
        update_ai_personal_rule(user, int(request.match_info['rule_id']), **data)
        return web.json_response({'ok': True, **init_payload(get_or_create_user(user['telegram_id'], user.get('full_name','Пользователь')))})
    except ValueError as e:
        return json_error(str(e), 400)

async def api_ai_rule_delete(request):
    try:
        user = auth_user(request)
        delete_ai_personal_rule(user, int(request.match_info['rule_id']))
        return web.json_response({'ok': True, **init_payload(get_or_create_user(user['telegram_id'], user.get('full_name','Пользователь')))})
    except ValueError as e:
        return json_error(str(e), 400)

async def api_budget_wizard(request):
    try:
        user = auth_user(request)
        if request.method == 'GET':
            return web.json_response({'ok': True, 'budget_wizard': get_budget_wizard_profile(fid(user))})
        data = await request.json()
        result = save_budget_wizard_profile(user, data)
        return web.json_response({'ok': True, 'budget_wizard': result, **init_payload(get_or_create_user(user['telegram_id'], user.get('full_name','Пользователь')))})
    except ValueError as e:
        return json_error(str(e), 400)

async def healthz(request):
    # Public enough for Railway healthcheck. Detailed DB error is shown only with HEALTHCHECK_SECRET.
    secret = request.query.get('secret', '')
    detailed = bool(HEALTHCHECK_SECRET and hmac.compare_digest(secret, HEALTHCHECK_SECRET))
    db_ok = False
    db_error = ''
    try:
        with engine.begin() as conn:
            conn.execute(text('SELECT 1'))
        db_ok = True
    except Exception as exc:
        db_error = str(exc)[:300]
        _capture_exception(exc)
    payload = {
        'ok': db_ok,
        'version': APP_VERSION,
        'database': 'postgresql' if str(DATABASE_URL).startswith('postgresql') else 'sqlite',
    }
    if detailed:
        payload['db_error'] = db_error
    return web.json_response(payload, status=200 if db_ok else 503)

async def privacy_policy(request):
    html_path = BASE_DIR / 'docs' / 'privacy_policy.html'
    if html_path.exists():
        return web.FileResponse(html_path)
    return web.Response(text='Privacy Policy', content_type='text/plain')

def create_web_app():
    app = web.Application(client_max_size=5*1024**2, middlewares=[security_headers_middleware, admin_security_middleware, maintenance_middleware, idempotency_middleware, error_middleware])
    app.router.add_get('/', index)
    app.router.add_get('/healthz', healthz)
    app.router.add_get('/privacy', privacy_policy)
    app.router.add_static('/webapp/', WEBAPP_DIR, show_index=False)
    app.router.add_static('/admin/static/', ADMIN_DIR, show_index=False)
    app.router.add_get('/admin', admin_index)
    app.router.add_post('/api/admin/login', api_admin_login)
    app.router.add_post('/api/admin/logout', api_admin_logout)
    app.router.add_get('/api/admin/stats', api_admin_stats)
    app.router.add_post('/api/admin/backup', api_admin_backup)
    app.router.add_get('/api/admin/backup/{name}', api_admin_backup_download)
    app.router.add_post('/api/admin/restore/request', api_admin_restore_request)
    app.router.add_post('/api/admin/restore', api_admin_backup_restore)
    app.router.add_post('/api/admin/users/role', api_admin_user_role)
    app.router.add_get('/api/admin/families/{family_id}', api_admin_family_detail)
    app.router.add_get('/api/admin/families/{family_id}/operations', api_admin_family_operations)
    app.router.add_post('/api/admin/users/block', api_admin_user_block)
    app.router.add_get('/api/admin/audit/export.xlsx', api_admin_audit_export)
    app.router.add_get('/api/init', api_init)
    app.router.add_delete('/api/account', api_delete_account)
    app.router.add_delete('/api/family', api_delete_family)
    app.router.add_get('/api/reports', api_reports)
    app.router.add_get('/api/calendar', api_financial_calendar)
    app.router.add_get('/api/mandatory-payments', api_mandatory_payments)
    app.router.add_get('/api/scheduled-payment-issues', api_scheduled_payment_issues)
    app.router.add_get('/api/month-end-money', api_month_end_money)
    app.router.add_post('/api/mandatory-payments/{schedule_id}/pay', api_mandatory_payment_pay)
    app.router.add_post('/api/mandatory-payments/{schedule_id}/mark-paid', api_mandatory_payment_mark_paid)
    app.router.add_get('/api/mandatory-payments/{schedule_id}/linkable-transactions', api_mandatory_payment_linkable_transactions)
    app.router.add_post('/api/mandatory-payments/{schedule_id}/link-existing', api_mandatory_payment_link_existing)
    app.router.add_post('/api/scheduled-payments/{schedule_id}/retry', api_scheduled_payment_retry)
    app.router.add_post('/api/scheduled-payments/{schedule_id}/disable-auto', api_scheduled_payment_disable_auto)
    app.router.add_post('/api/scheduled-payment-issues/{issue_id}/resolve', api_scheduled_payment_issue_resolve)
    app.router.add_get('/api/ai-rules', api_ai_rules)
    app.router.add_post('/api/ai-rules', api_ai_rules)
    app.router.add_put('/api/ai-rules/{rule_id}', api_ai_rule_update)
    app.router.add_delete('/api/ai-rules/{rule_id}', api_ai_rule_delete)
    app.router.add_get('/api/budget-wizard', api_budget_wizard)
    app.router.add_post('/api/budget-wizard', api_budget_wizard)
    app.router.add_get('/api/ai-analysis', api_ai_analysis)
    app.router.add_get('/api/charts', api_chart_pack)
    app.router.add_get('/api/forecast', api_forecast)
    app.router.add_get('/api/report.pdf', api_monthly_pdf)
    app.router.add_get('/api/notification-settings', api_notification_settings)
    app.router.add_post('/api/notification-settings', api_notification_settings)
    app.router.add_get('/api/export.xlsx', api_export)
    app.router.add_get('/api/operations', api_operations)
    app.router.add_get('/api/operations/{tx_id}/history', api_operation_history)
    app.router.add_get('/api/audit-logs', api_audit_logs)
    app.router.add_post('/api/undo', api_undo)
    app.router.add_get('/api/wallets', api_wallets)
    app.router.add_post('/api/wallets', api_wallets)
    app.router.add_put('/api/wallets/{wallet_id}', api_wallet_update)
    app.router.add_delete('/api/wallets/{wallet_id}', api_wallet_delete)
    app.router.add_get('/api/categories', api_categories)
    app.router.add_post('/api/categories', api_categories)
    app.router.add_put('/api/categories/{category_id}', api_category_update)
    app.router.add_delete('/api/categories/{category_id}', api_category_delete)
    app.router.add_post('/api/transactions', api_transaction)
    app.router.add_put('/api/transactions/{tx_id}', api_transaction_edit)
    app.router.add_delete('/api/transactions/{tx_id}', api_transaction_delete)
    app.router.add_post('/api/transfers', api_transfer)
    app.router.add_delete('/api/transfers/{transfer_id}', api_transfer_delete)
    app.router.add_post('/api/family/join', api_join_family)
    app.router.add_post('/api/family/member/role', api_family_member_role)
    app.router.add_post('/api/family/member/remove', api_family_member_remove)
    app.router.add_get('/api/family/member/{member_id}/permissions', api_family_member_permissions)
    app.router.add_post('/api/family/member/{member_id}/permissions', api_family_member_permissions)
    app.router.add_post('/api/rates', api_rates)
    app.router.add_post('/api/debts', api_debt)
    app.router.add_post('/api/debts/pay', api_pay_debt)
    app.router.add_post('/api/goals', api_goal)
    app.router.add_post('/api/goals/add', api_goal_add)
    app.router.add_post('/api/budgets', api_budget)
    app.router.add_get('/api/scheduled-payments', api_scheduled_payments)
    app.router.add_post('/api/scheduled-payments', api_scheduled_payments)
    app.router.add_put('/api/scheduled-payments/{schedule_id}', api_scheduled_payment_update)
    app.router.add_delete('/api/scheduled-payments/{schedule_id}', api_scheduled_payment_delete)
    app.router.add_get('/api/financial-plan', api_financial_plan)
    app.router.add_post('/api/financial-plan', api_financial_plan)
    app.router.add_put('/api/financial-plan/{item_id}', api_financial_plan_update)
    app.router.add_delete('/api/financial-plan/{item_id}', api_financial_plan_delete)
    return app
